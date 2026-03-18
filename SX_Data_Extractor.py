

# ── CELL 1 — Imports & API key ───────────────────────────────────────────────
import anthropic
import base64, json, logging, math, os, re, sys, textwrap, io
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("agent")

# ── Load API key from file ────────────────────────────────────────────────────
# Create a plain-text file containing only your Anthropic API key (sk-ant-...)
# Get one at: https://console.anthropic.com/
API_KEY_PATH   = r"path to txt file containing the API key goes here"
ANTHROPIC_KEY  = Path(API_KEY_PATH).read_text(encoding="utf-8").strip()

# ── Create Anthropic client ───────────────────────────────────────────────────
client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)

# ── Model selection ───────────────────────────────────────────────────────────
# claude-opus-4-6   — most capable, best for complex extraction
# claude-sonnet-4-6 — faster and cheaper
MODEL = "claude-opus-4-6"

# ── PDF path — edit this to your file ────────────────────────────────────────
PDF_PATH = r"path to the pdf file goes here"

# Output will be saved next to the PDF
OUTPUT_PATH = str(Path(PDF_PATH).parent / (Path(PDF_PATH).stem + "_extracted.xlsx"))

print(f"PDF    : {PDF_PATH}")
print(f"Output : {OUTPUT_PATH}")
print(f"Model  : {MODEL}")
print("API key loaded ✓")

# Quick connectivity check
_test = client.messages.create(
    model=MODEL,
    max_tokens=16,
    messages=[{"role": "user", "content": "Hello, Claude"}],
)
print(f"Connection test ✓  ({_test.content[0].text.strip()})")


# ── CELL 2 — Lookup tables ───────────────────────────────────────────────────

CATION_TABLE: dict[str, list[tuple[int, float]]] = {
    "Ag": [(1, 1.29)],  "Al": [(3, 0.675)], "Am": [(3, 1.115)], "As": [(5, 0.60)],
    "Au": [(3, 0.99)],  "Ba": [(2, 1.49)],  "Bi": [(3, 1.17)],  "Bk": [(3, 1.10)],
    "Ca": [(2, 1.14)],  "Cd": [(2, 1.09)],  "Ce": [(3, 1.15)],  "Cf": [(3, 1.09)],
    "Cm": [(3, 1.11)],  "Co": [(2, 0.885)], "Cr": [(3, 0.755), (6, 0.58)],
    "Cs": [(1, 1.81)],  "Cu": [(1, 0.91), (2, 0.87)],
    "Dy": [(3, 1.052)], "Er": [(3, 1.030)], "Eu": [(3, 1.087)],
    "Fe": [(2, 0.920), (3, 0.785)],
    "Ga": [(3, 0.76)],  "Gd": [(3, 1.078)], "Ge": [(4, 0.67)],  "H":  [(1, -0.24)],
    "Hf": [(4, 0.85)],  "Hg": [(2, 1.16)],  "Ho": [(3, 1.041)], "In": [(3, 0.94)],
    "Ir": [(4, 0.765)], "La": [(3, 1.172)], "Lu": [(3, 1.001)], "Mn": [(2, 0.97)],
    "Mo": [(6, 0.73)],  "Nb": [(5, 0.78)],  "Nd": [(3, 1.12)],  "Ni": [(2, 0.83)],
    "Np": [(4, 1.01), (6, 0.86)],
    "Os": [(4, 0.77)],  "Pa": [(5, 0.92)],  "Pb": [(2, 1.33)],  "Pd": [(2, 1.00)],
    "Pm": [(3, 1.11)],  "Po": [(4, 1.08)],  "Pr": [(3, 1.13)],  "Pt": [(4, 0.765)],
    "Pu": [(3, 1.14), (6, 0.85)],
    "Rb": [(1, 1.66)],  "Re": [(7, 0.67)],  "Rh": [(3, 0.805)], "Ru": [(3, 0.82)],
    "Sb": [(3, 0.90)],  "Sc": [(3, 0.885)], "Se": [(4, 0.64)],  "Sm": [(3, 1.098)],
    "Sn": [(4, 0.83)],  "Sr": [(2, 1.32)],  "Ta": [(5, 0.78)],  "Tb": [(3, 1.063)],
    "Tc": [(7, 0.70)],  "Te": [(4, 1.11)],  "Th": [(4, 1.08)],  "Ti": [(4, 0.745)],
    "Tl": [(3, 1.025)], "Tm": [(3, 1.020)],
    "U":  [(4, 1.03), (6, 0.87)],
    "V":  [(5, 0.68)],  "W":  [(6, 0.74)],  "Y":  [(3, 1.04)],  "Yb": [(3, 1.008)],
    "Zn": [(2, 0.88)],  "Zr": [(4, 0.80)],
}

ANION_TABLE: dict[str, float] = {
    "bromide": 1.82, "chloride": 1.67, "fluoride": 1.19,
    "nitrate": 0.27, "perchlorate": 0.41, "phosphate": 0.52, "sulfate": 0.43,
}

SOLVENT_TABLE: dict[str, float] = {
    "kerosene": 1.80,            "heptane": 1.92,
    "octane": 1.95,              "dodecane": 2.01,
    "carbon tetrachloride": 2.24,"ccl4": 2.24,
    "p-xylene": 2.27,            "xylene": 2.27,
    "benzene": 2.28,             "toluene": 2.38,
    "dibutyl ether": 3.08,       "dibutylether": 3.08,
    "chloroform": 4.80,          "2-octanol": 8.13,
    "tri-butyl phosphate": 8.34, "tributyl phosphate": 8.34,
    "tbp": 8.34,                 "1-octanol": 10.30,
}

TBP_MW      = 266.32   # g/mol
TBP_DENSITY = 0.9727   # g/mL

COLUMNS = [
    "Source PDF", "Figure / Table", "Metal (cation)",
    "Acid concentration (mol/L)", "Size cation extracted (Å)",
    "Oxidation state cation",
    "Initial aqueous cation concentration (g/L)",
    "Anion", "Anion concentration (mol/L)",
    "TBP concentration (%v/v)", "TBP concentration (mol/L)",
    "Size Anion (Å)", "Dielectric constant solvent", "Solvent",
    "O/A ratio", "Distribution ratio D", "Extraction efficiency E (%)",
    "Partition coefficient (raw)", "Notes / confidence",
]

print("Lookup tables loaded ✓")


# ── CELL 3 — Claude helpers ──────────────────────────────────────────────────

def _chat(prompt: str, system: str = "") -> str:
    """Text-only Claude call."""
    kwargs = dict(
        model=MODEL,
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    if system:
        kwargs["system"] = system
    message = client.messages.create(**kwargs)
    return message.content[0].text


def _image_to_base64(image_path: str) -> tuple[str, str]:
    """Convert an image file to base64 and return (b64_string, media_type)."""
    img = PILImage.open(image_path)
    buf = io.BytesIO()
    fmt = (img.format or "PNG").upper()
    if fmt not in ("PNG", "JPEG", "GIF", "WEBP"):
        fmt = "PNG"
    img.save(buf, format=fmt)
    b64       = base64.standard_b64encode(buf.getvalue()).decode("utf-8")
    mime_type = f"image/{fmt.lower()}"
    return b64, mime_type


def _vision(prompt: str, image_path: str) -> str:
    """Multimodal Claude call — image + text prompt."""
    b64, media_type = _image_to_base64(image_path)
    message = client.messages.create(
        model=MODEL,
        max_tokens=4096,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type":       "base64",
                            "media_type": media_type,
                            "data":       b64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )
    return message.content[0].text


def _parse_json(raw: str) -> object:
    """Strip markdown fences and parse first JSON structure found."""
    clean = re.sub(r"```(?:json)?|```", "", raw).strip()
    m = re.search(r"[{\[]", clean)
    if m:
        clean = clean[m.start():]
    return json.loads(clean)

print("Claude helpers ready ✓")


# ── CELL 4 — PDF parser ──────────────────────────────────────────────────────

@dataclass
class ParsedPDF:
    full_text: str = ""
    tables:    list = field(default_factory=list)
    figures:   list = field(default_factory=list)


def parse_pdf(pdf_path: str) -> ParsedPDF:
    """Extract text + tables (pdfplumber) and figure images (PyMuPDF)."""
    work_dir = Path(pdf_path).parent / (Path(pdf_path).stem + "_agent_tmp")
    work_dir.mkdir(parents=True, exist_ok=True)

    result     = ParsedPDF()
    pages_text = []

    # Text and tables
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            pages_text.append(page.extract_text() or "")
            for raw_tbl in page.extract_tables():
                if raw_tbl and len(raw_tbl) >= 2:
                    result.tables.append({"label": "", "caption": "", "data": raw_tbl})

    result.full_text = "\n".join(pages_text)

    # Figure images via PyMuPDF
    try:
        import fitz
        doc     = fitz.open(pdf_path)
        fig_idx = 0
        for pg in range(len(doc)):
            for img_info in doc[pg].get_images(full=True):
                xref = img_info[0]
                img  = doc.extract_image(xref)
                if img["width"] < 150 or img["height"] < 150 or len(img["image"]) < 8_000:
                    continue
                fig_idx += 1
                out_path = work_dir / f"fig_p{pg+1}_{fig_idx}.{img['ext']}"
                out_path.write_bytes(img["image"])
                result.figures.append({
                    "label":      f"Figure {fig_idx}",
                    "caption":    "",
                    "image_path": str(out_path),
                    "page":       pg + 1,
                })
        doc.close()
    except ImportError:
        log.warning("PyMuPDF not installed — figure extraction skipped.\n"
                    "Install with:  pip install PyMuPDF")

    # Assign captions from text (figures and tables)
    cap_re = re.compile(
        r"(?P<lbl>(?:Fig(?:ure)?|Table)\.?\s*\d+[a-z]?)[.:\s]+"
        r"(?P<txt>[^\n]{10,400})", re.IGNORECASE)
    fig_caps, tbl_caps = [], []
    for m in cap_re.finditer(result.full_text):
        lbl, txt = m.group("lbl").strip(), m.group("txt").strip()
        (fig_caps if re.match(r"fig", lbl, re.I) else tbl_caps).append((lbl, txt))
    for i, fig in enumerate(result.figures):
        if i < len(fig_caps):
            fig["label"], fig["caption"] = fig_caps[i]
    for i, tbl in enumerate(result.tables):
        if i < len(tbl_caps):
            tbl["label"], tbl["caption"] = tbl_caps[i]

    log.info(f"PDF parsed: {len(pages_text)} pages | "
             f"{len(result.tables)} tables | {len(result.figures)} figures")
    return result


# ── CELL 5 — Text agent ──────────────────────────────────────────────────────

_TEXT_SYS = (
    "You are an expert chemist specialising in TBP solvent extraction. "
    "Extract structured data from paper text. "
    "Respond with valid JSON only — no prose, no markdown fences."
)

_TEXT_PROMPT = """\
Read this excerpt from a solvent extraction paper and extract the items below.

EXTRACT:
1. metal_symbol      — element symbol e.g. "U", "Au", "Fe". One only.
2. metal_name        — full English name e.g. "gold".
3. oxidation_state   — integer. Use the specific state mentioned. If ambiguous,
                       use the most common state for this metal.
4. anion             — counter-anion in the aqueous phase. Use ONLY one of:
                         bromide | chloride | fluoride | nitrate |
                         perchlorate | phosphate | sulfate
5. solvent           — organic diluent name. Match to:
                         kerosene | heptane | octane | dodecane |
                         carbon tetrachloride | p-xylene | xylene | benzene |
                         toluene | dibutyl ether | chloroform |
                         2-octanol | 1-octanol | tbp
                       (use "tbp" ONLY if TBP is used pure with no diluent)
6. oa_ratio          — O/A volume ratio as a float. null if not stated.
7. initial_conc_g_l  — initial metal concentration in the AQUEOUS phase before
                       extraction, in g/L. null if absent.
8. global_acid_mol_l — single fixed acid conc (mol/L) used in ALL experiments,
                       or null if it varies.
9. global_tbp_pct    — single fixed TBP vol% used in ALL experiments, or null.

10. equations — ANY algebraic equations that can be used to derive D, E%,
    partition coefficient, or distribution ratio. This includes:
      - Direct equations: log(D) = a + b·log[TBP], D = K·[TBP]^n
      - Extraction equilibrium constants: Kex expressions
      - Partition coefficient equations: P = ...
      - Any equation from which D or E can be inferred given typical conditions

    Examples of valid equation formats:
      "log D = 2.1 + 3.0 * log[TBP]"              (slope-intercept vs TBP)
      "log D = -1.5 + 2.0 * log[H+] + 1.0 * log[NO3-]"  (vs acid components)
      "D = Kex * [TBP]^2 * [H+]^2 / [H2O]^2"     (extraction constant form)
      "log Kex = 4.5"                               (constant, use to infer D)
      "P = 10^(0.8 * pH - 1.2)"                    (partition coeff vs pH)
      "E% = 100 * D / (D + 1)"                     (E from D, O/A=1)

    Include x_min/x_max if the valid range is clearly stated or visible in
    the figure; otherwise set both to null — the equation will still be used.

    Each equation object:
    {{
      "expr":    "log(D) = 3.52 + 2.83*log(TBP)",
      "lhs":     "log_D",      // "D"|"log_D"|"E"|"log_E"|"Kex"|"log_Kex"|"P"|"log_P"
      "x_var":   "TBP",        // "TBP"|"acid"|"H+"|"NO3-"|"pH"|"none"
      "x_unit":  "mol/L",      // "mol/L"|"vol%"|"pH units"|"none"
      "x_min":   null,         // lower bound of valid range, or null
      "x_max":   null,         // upper bound of valid range, or null
      "fixed": {{
        "acid_mol_l": 3.0,
        "tbp_pct":    null,
        "oa_ratio":   null
      }},
      "interpretation": "brief note on how to derive D from this equation",
      "source": "Fig 4 regression"
    }}

11. text_points — explicit (acid, TBP, D, partition coefficient or E%) values
    stated in running text (not in tables/figures):
    {{
      "acid_mol_l":          <number|null>,
      "tbp_pct":             <number|null>,
      "tbp_mol_l":           <number|null>,
      "D":                   <number|null>,
      "partition_coeff":     <number|null>,
      "E_pct":               <number 0-100|null>,
      "notes":               "stated in Section 3.1"
    }}

Return a SINGLE JSON object with keys:
metal_symbol, metal_name, oxidation_state, anion, solvent, oa_ratio,
initial_conc_g_l, global_acid_mol_l, global_tbp_pct, equations, text_points

TEXT:
{text}
"""

def run_text_agent(parsed: ParsedPDF) -> dict:
    log.info("TEXT AGENT — extracting metadata, equations...")
    prompt = _TEXT_PROMPT.format(text=parsed.full_text[:14_000])
    raw    = _chat(prompt, system=_TEXT_SYS)
    try:
        meta = _parse_json(raw)
    except Exception as e:
        log.error(f"Text agent JSON error: {e}\nRaw output:\n{raw[:600]}")
        meta = {}
    log.info(f"  metal={meta.get('metal_symbol')}  "
             f"ox={meta.get('oxidation_state')}  "
             f"anion={meta.get('anion')}  "
             f"solvent={meta.get('solvent')}  "
             f"O/A={meta.get('oa_ratio')}  "
             f"equations={len(meta.get('equations', []))}")
    return meta


# ── CELL 6 — Table agent ─────────────────────────────────────────────────────

_TABLE_SYS = (
    "You are an expert data extractor for solvent extraction chemistry tables. "
    "Return only valid JSON. No prose, no markdown."
)

_TABLE_PROMPT = """\
Table from a TBP solvent extraction paper.
Extract every data row. Target values: distribution ratio D, partition
coefficient (P), or extraction efficiency E (%).

Note: "Partition coefficient" and "distribution ratio" are equivalent terms
and should both be captured. Report the raw value in "partition_coeff" if the
paper uses that term, and also propagate it to "D" since they are equivalent.

Rules:
  - If D is given as log(D), convert: D = 10^(logD)
  - If E is given as fraction 0–1, convert to percent (* 100)
  - Blank/dash cells → null
  - "Initial metal concentration" means concentration in the aqueous phase
    BEFORE extraction

Return a JSON array — one object per row:
{{
  "acid_mol_l":          <number|null>,
  "tbp_pct":             <TBP vol%|null>,
  "tbp_mol_l":           <TBP mol/L|null>,
  "D":                   <distribution ratio or partition coeff, number|null>,
  "partition_coeff":     <raw partition coeff value if named as such|null>,
  "E_pct":               <number 0-100|null>,
  "metal_conc_g_l":      <initial aqueous metal conc before extraction, g/L|null>,
  "oa_ratio":            <number|null>,
  "notes":               "<temperature or other qualifiers>"
}}

Table label:   {label}
Table caption: {caption}
Table (pipe-separated rows):
{table_str}
"""

def run_table_agent(parsed: ParsedPDF) -> list[dict]:
    rows = []
    for tbl in parsed.tables:
        if not tbl["data"] or len(tbl["data"]) < 2:
            continue
        table_str = "\n".join(
            " | ".join(str(c or "").strip() for c in row)
            for row in tbl["data"]
        )
        prompt = _TABLE_PROMPT.format(
            label=tbl.get("label", "unknown"),
            caption=tbl.get("caption", ""),
            table_str=table_str,
        )
        log.info(f"TABLE AGENT — {tbl.get('label', 'table')}...")
        raw = _chat(prompt, system=_TABLE_SYS)
        try:
            parsed_rows = _parse_json(raw)
            if isinstance(parsed_rows, list):
                for r in parsed_rows:
                    r["_source"] = tbl.get("label", "Table")
                rows.extend(parsed_rows)
                log.info(f"  → {len(parsed_rows)} rows extracted")
        except Exception as e:
            log.warning(f"  Table parse error: {e}")
    return rows


# ── CELL 7 — Figure agent (vision) ──────────────────────────────────────────

_FIG_PROMPT = """\
This image is from a TBP solvent extraction paper.
Figure label:   {label}
Figure caption: {caption}

STEP 1 — Is this a data plot showing distribution ratio D, partition coefficient,
or extraction efficiency E (as a function of any variable)?
Relevant variables include: TBP concentration, acid concentration, pH,
initial metal concentration, O/A ratio, time, or any other experimental parameter.
If NOT a data plot (schematic, photo, flow diagram, structure), return:
  {{"is_data_plot": false}}

STEP 2 — If YES, carefully read the caption above for context. The caption may
tell you:
  - Which metals are shown (if multiple series)
  - The fixed experimental conditions (e.g. "3 mol/L HNO3", "30% TBP")
  - What the different symbols/colours represent
  - Whether y-axis is D, log D, E%, partition coefficient, etc.

Use that caption context to correctly label each data series.

Extract every data series:
{{
  "is_data_plot": true,
  "x_axis_label": "<label with units>",
  "y_axis_label": "<label with units>",
  "y_is_log":     <true if y-axis is logarithmic>,
  "x_is_log":     <true if x-axis is logarithmic>,
  "y_quantity":   "D"|"log_D"|"E_pct"|"partition_coeff"|"log_partition_coeff"|"unknown",
  "x_quantity":   "TBP_pct"|"TBP_mol_L"|"acid_mol_L"|"pH"|"metal_conc_g_L"|"other",
  "series": [
    {{
      "name": "<legend label — include metal name/symbol if multiple metals shown>",
      "metal_override": "<element symbol if this series is a DIFFERENT metal than
                          the main paper subject, else null>",
      "fixed_conditions": {{
        "acid_mol_l":     <number|null>,
        "tbp_pct":        <number|null>,
        "metal_conc_g_l": <initial aqueous conc before extraction, g/L|null>,
        "oa_ratio":       <number|null>,
        "pH":             <number|null>
      }},
      "points": [[x1,y1], [x2,y2], ...]
    }}
  ]
}}

IMPORTANT RULES:
  - Read log-scale axes correctly (do NOT linearise — return the raw axis value)
  - If y-axis is log(D) or log scale, return RAW axis value and set y_is_log=true
  - If E shown as fraction 0–1, multiply by 100 before returning
  - If multiple metals appear in one figure, create a separate series for each
    with "metal_override" set to the element symbol
  - Fixed conditions should be taken from the caption text first, then the legend
  - "metal_conc_g_l" always refers to the initial concentration in the AQUEOUS
    phase before extraction

Return ONLY the JSON object.
"""

def _pct_to_mol(pct) -> float | None:
    if pct is None: return None
    return round(float(pct) * 10 * TBP_DENSITY / TBP_MW, 5)

def _mol_to_pct(mol) -> float | None:
    if mol is None: return None
    return round(float(mol) * TBP_MW / (TBP_DENSITY * 10), 4)


def run_figure_agent(parsed: ParsedPDF) -> list[dict]:
    rows = []
    for fig in parsed.figures:
        img_path = fig.get("image_path", "")
        if not Path(img_path).exists():
            continue
        prompt = _FIG_PROMPT.format(
            label=fig.get("label", ""),
            caption=fig.get("caption", ""),
        )
        log.info(f"FIGURE AGENT — {fig.get('label', Path(img_path).name)}...")
        try:
            raw    = _vision(prompt, img_path)
            result = _parse_json(raw)
        except Exception as e:
            log.warning(f"  Vision error: {e}")
            continue

        if not result.get("is_data_plot"):
            log.info("  → not a data plot, skipped")
            continue

        x_lbl     = result.get("x_axis_label", "").lower()
        y_lbl     = result.get("y_axis_label", "").lower()
        y_log     = result.get("y_is_log", False)
        y_qty     = result.get("y_quantity", "unknown").lower()
        x_qty     = result.get("x_quantity", "").lower()

        # Determine x-axis type from explicit field or label fallback
        x_is_tbp  = "tbp" in x_qty or "tbp" in x_lbl or "extractant" in x_lbl
        x_is_acid = ("acid" in x_qty or any(
            a in x_lbl for a in ("acid","hno3","hcl","hclo4","h2so4","hbr")))
        x_is_pH   = "ph" in x_qty or x_lbl.strip().startswith("ph")
        x_is_metal= "metal" in x_qty or "conc" in x_lbl

        # Determine y-axis meaning
        y_is_D    = y_qty in ("d","log_d") or "distribut" in y_lbl or re.search(r"\bD\b", y_lbl) is not None
        y_is_E    = y_qty == "e_pct" or any(e in y_lbl for e in ("extract","effici","%e","e %","e(%)"))
        y_is_P    = y_qty in ("partition_coeff","log_partition_coeff") or "partition" in y_lbl

        fig_rows_before = len(rows)
        for series in result.get("series", []):
            fixed         = series.get("fixed_conditions", {})
            metal_override= series.get("metal_override")  # different metal in multi-metal fig

            for pt in series.get("points", []):
                if len(pt) < 2: continue
                xv, yv = float(pt[0]), float(pt[1])

                # De-log the y value if the axis was logarithmic
                if y_log:
                    yv = 10 ** yv

                # Assign D / E / partition_coeff based on y-axis type
                d_val    = None
                e_val    = None
                part_val = None

                if y_is_D or (not y_is_E and not y_is_P):
                    d_val = yv
                if y_is_E:
                    e_val = yv
                if y_is_P:
                    part_val = yv
                    d_val    = yv   # treat as equivalent to D

                # Map x-axis to the right column
                if x_is_tbp:
                    if "mol" in x_lbl:
                        tbp_mol, tbp_pct = xv, _mol_to_pct(xv)
                    else:
                        tbp_pct, tbp_mol = xv, _pct_to_mol(xv)
                    acid = fixed.get("acid_mol_l")
                    pH   = fixed.get("pH")
                else:
                    acid    = xv if x_is_acid else fixed.get("acid_mol_l")
                    pH      = xv if x_is_pH   else fixed.get("pH")
                    tbp_pct = fixed.get("tbp_pct")
                    tbp_mol = _pct_to_mol(tbp_pct) if tbp_pct else None

                rows.append({
                    "_source":        fig.get("label", "Figure"),
                    "_metal_override":metal_override,   # handled in assembler
                    "acid_mol_l":     acid,
                    "pH":             pH,
                    "tbp_pct":        tbp_pct,
                    "tbp_mol_l":      tbp_mol,
                    "D":              d_val,
                    "partition_coeff":part_val,
                    "E_pct":          e_val,
                    "metal_conc_g_l": fixed.get("metal_conc_g_l"),
                    "oa_ratio":       fixed.get("oa_ratio"),
                    "notes":          series.get("name", ""),
                })
        log.info(f"  → {len(rows) - fig_rows_before} rows from this figure")
    return rows


# ── CELL 8 — Equation evaluator ─────────────────────────────────────────────

# Safe math context for eval
_MATH_ENV = {
    "__builtins__": {},
    "log10": math.log10, "log": math.log10, "ln": math.log,
    "exp":   math.exp,   "sqrt": math.sqrt, "pow": math.pow,
    "pi":    math.pi,
}

# How to derive D from various LHS types
_LHS_TO_D = {
    "d":          lambda y: y,
    "log_d":      lambda y: 10 ** y,
    "log d":      lambda y: 10 ** y,
    "e":          None,   # handled separately
    "log_e":      None,
    "e_pct":      None,
    "partition_coeff":     lambda y: y,       # treat as equivalent to D
    "log_partition_coeff": lambda y: 10 ** y,
    "p":          lambda y: y,
    "log_p":      lambda y: 10 ** y,
    "kex":        None,   # need interpretation
    "log_kex":    None,
}

_LHS_TO_E = {
    "e":      lambda y: y * 100 if y <= 1 else y,
    "log_e":  lambda y: (10 ** y) * 100,
    "e_pct":  lambda y: y,
}


def _build_rhs(expr: str) -> str:
    """Normalise an equation expression string for eval."""
    rhs = expr.split("=", 1)[-1].strip()
    # Replace common symbolic names with eval-friendly forms
    replacements = [
        (r"log\(TBP\)",        "log10(x)"),
        (r"log\[TBP\]",        "log10(x)"),
        (r"log10\(TBP\)",      "log10(x)"),
        (r"log\(acid\)",       "log10(x)"),
        (r"log\[acid\]",       "log10(x)"),
        (r"log\(HNO3\)",       "log10(x)"),
        (r"log\[HNO3\]",       "log10(x)"),
        (r"log\(HCl\)",        "log10(x)"),
        (r"log\[HCl\]",        "log10(x)"),
        (r"log\(H\+\)",        "log10(x)"),
        (r"log\[H\+\]",        "log10(x)"),
        (r"log\(NO3-\)",       "log10(x)"),
        (r"log\[NO3-\]",       "log10(x)"),
        (r"log10\(acid\)",     "log10(x)"),
        (r"log10\(H\+\)",      "log10(x)"),
        (r"\bTBP\b",           "x"),
        (r"\bacid\b",          "x"),
        (r"\bHNO3\b",          "x"),
        (r"\bHCl\b",           "x"),
        (r"\bpH\b",            "x"),
        (r"\^",                "**"),
    ]
    for pattern, repl in replacements:
        rhs = re.sub(pattern, repl, rhs)
    return rhs


def _default_range(x_var: str, x_unit: str) -> tuple[float, float, int]:
    """Return (x_min, x_max, n_points) defaults when no range is given."""
    x_var_l = x_var.lower()
    if "tbp" in x_var_l:
        if "mol" in x_unit.lower():
            return 0.05, 1.5, 10
        return 5.0, 100.0, 10
    if "ph" in x_var_l:
        return 0.0, 14.0, 15
    # acid concentration default
    return 0.1, 10.0, 10


def run_equation_agent(equations: list[dict], meta: dict) -> list[dict]:
    """Evaluate algebraic D/E equations over their stated (or inferred) range."""
    rows = []
    for eq in equations:
        x_var  = eq.get("x_var", "none").lower()
        x_unit = eq.get("x_unit", "mol/L").lower()
        lhs    = eq.get("lhs", "log_D").lower().replace(" ", "_")
        fixed  = eq.get("fixed", {})
        interp = eq.get("interpretation", "")

        x_min = eq.get("x_min")
        x_max = eq.get("x_max")

        # If no range is supplied, use sensible defaults instead of skipping
        if x_min is None or x_max is None:
            if x_var == "none":
                # Constant equation — evaluate once at x=1 (dummy)
                x_min, x_max, n_pts = 1.0, 1.0, 1
            else:
                x_min, x_max, n_pts = _default_range(x_var, x_unit)
                log.info(f"  Equation '{eq.get('expr')}' — no range given, "
                         f"using default {x_min}–{x_max}")
        else:
            n_pts = 10

        rhs = _build_rhs(eq.get("expr", ""))

        acid_fixed = fixed.get("acid_mol_l") or meta.get("global_acid_mol_l")
        tbp_fixed  = fixed.get("tbp_pct")    or meta.get("global_tbp_pct")

        xs = ([x_min] if n_pts == 1
              else [x_min + i * (x_max - x_min) / (n_pts - 1) for i in range(n_pts)])

        for x in xs:
            try:
                env = dict(_MATH_ENV)
                env["x"] = x
                y = eval(rhs, env)
            except Exception as e:
                log.debug(f"Equation eval error for '{rhs}' at x={x}: {e}")
                continue

            # Derive D and E
            d_val = None
            e_val = None
            part_val = None

            d_fn = _LHS_TO_D.get(lhs)
            e_fn = _LHS_TO_E.get(lhs)

            if d_fn:
                try:
                    d_val = d_fn(y)
                except Exception:
                    pass
            if e_fn:
                try:
                    e_val = e_fn(y)
                except Exception:
                    pass

            # For Kex / extraction constants: attempt to infer D
            # The model's "interpretation" field guides this
            if lhs in ("kex", "log_kex") and d_val is None:
                # Common form: Kex = D / ([TBP]^n * [H+]^m)
                # Without further symbolic parsing, store as note
                d_val = None
                e_val = None
                log.debug(f"  Kex equation — D inference requires "
                          f"additional context: {interp}")

            if lhs in ("p", "log_p", "partition_coeff", "log_partition_coeff"):
                part_val = d_val  # same value, label differently

            # Map x to the right column
            if "tbp" in x_var:
                if "mol" in x_unit:
                    tbp_mol, tbp_pct = x, _mol_to_pct(x)
                else:
                    tbp_pct, tbp_mol = x, _pct_to_mol(x)
                acid = acid_fixed
            elif "ph" in x_var:
                acid    = None
                tbp_pct = tbp_fixed
                tbp_mol = _pct_to_mol(tbp_pct) if tbp_pct else None
            elif x_var == "none":
                acid    = acid_fixed
                tbp_pct = tbp_fixed
                tbp_mol = _pct_to_mol(tbp_pct) if tbp_pct else None
            else:
                acid    = x
                tbp_pct = tbp_fixed
                tbp_mol = _pct_to_mol(tbp_pct) if tbp_pct else None

            rows.append({
                "_source":        eq.get("source", "equation"),
                "_metal_override":None,
                "acid_mol_l":     acid,
                "pH":             x if "ph" in x_var else None,
                "tbp_pct":        tbp_pct,
                "tbp_mol_l":      tbp_mol,
                "D":              d_val,
                "partition_coeff":part_val,
                "E_pct":          e_val,
                "metal_conc_g_l": None,
                "oa_ratio":       fixed.get("oa_ratio"),
                "notes":          (f"Equation: {eq.get('expr')}"
                                   + (f"  [{interp}]" if interp else "")
                                   + f"  (x_range: {x_min}–{x_max} {x_unit})"),
            })

    log.info(f"EQUATION AGENT — {len(rows)} rows from {len(equations)} equations")
    return rows


# ── CELL 9 — Assembler ───────────────────────────────────────────────────────

def _coerce(v):
    if v is None or v == "": return None
    try:    return round(float(v), 5)
    except: return None


def _get_metal_props(sym: str, ox):
    """Return (cat_ox, cat_radius) for a given element symbol and oxidation state."""
    entries = CATION_TABLE.get(sym, [])
    if not entries:
        return None, None
    if ox is not None:
        try:
            ox_int = int(ox)
            cat_radius = next((r for s, r in entries if s == ox_int), entries[0][1])
            return ox_int, cat_radius
        except (TypeError, ValueError):
            pass
    return entries[0]


def assemble(raw_rows: list[dict], meta: dict, pdf_name: str) -> pd.DataFrame:
    # Primary metal from text agent
    sym_primary = (meta.get("metal_symbol") or "").strip()
    ox_primary  = meta.get("oxidation_state")
    cat_ox_p, cat_radius_p = _get_metal_props(sym_primary, ox_primary)

    anion      = (meta.get("anion")   or "").lower().strip()
    anion_size = ANION_TABLE.get(anion)
    solvent    = (meta.get("solvent") or "").lower().strip()
    diel       = SOLVENT_TABLE.get(solvent)
    oa_global  = meta.get("oa_ratio")
    conc_global= meta.get("initial_conc_g_l")

    final = []
    for r in raw_rows:
        # Handle metal override (multi-metal figures)
        metal_override = r.get("_metal_override")
        if metal_override and metal_override.strip():
            sym        = metal_override.strip()
            cat_ox, cat_radius = _get_metal_props(sym, None)
        else:
            sym        = sym_primary
            cat_ox     = cat_ox_p
            cat_radius = cat_radius_p

        acid  = _coerce(r.get("acid_mol_l") or meta.get("global_acid_mol_l"))
        tbp_p = _coerce(r.get("tbp_pct")    or meta.get("global_tbp_pct"))
        tbp_m = _coerce(r.get("tbp_mol_l"))
        if tbp_p is not None and tbp_m is None: tbp_m = _coerce(_pct_to_mol(tbp_p))
        if tbp_m is not None and tbp_p is None: tbp_p = _coerce(_mol_to_pct(tbp_m))

        oa     = _coerce(r.get("oa_ratio") or oa_global)
        # initial_conc_g_l = concentration in aqueous phase BEFORE extraction
        conc   = _coerce(r.get("metal_conc_g_l") or conc_global)
        d_val  = _coerce(r.get("D"))
        e_val  = _coerce(r.get("E_pct"))
        p_val  = _coerce(r.get("partition_coeff"))

        # If partition_coeff is present and D is not, use it as D
        if d_val is None and p_val is not None:
            d_val = p_val

        oa_num = float(oa) if oa is not None else 1.0
        # Cross-compute D <-> E
        if d_val is None and e_val is not None:
            try:
                ef = float(e_val) / 100
                d_val = _coerce(ef * oa_num / (1 - ef)) if 0 < ef < 1 else None
            except: pass
        if e_val is None and d_val is not None:
            try:
                d_num = float(d_val)
                e_val = _coerce(100 * d_num / (d_num + oa_num))
            except: pass

        final.append({
            "Source PDF":                               pdf_name,
            "Figure / Table":                           r.get("_source", ""),
            "Metal (cation)":                           sym,
            "Acid concentration (mol/L)":               acid,
            "Size cation extracted (Å)":                cat_radius,
            "Oxidation state cation":                   cat_ox,
            "Initial aqueous cation concentration (g/L)": conc,
            "Anion":                                    anion,
            "Anion concentration (mol/L)":              acid,
            "TBP concentration (%v/v)":                 tbp_p,
            "TBP concentration (mol/L)":                tbp_m,
            "Size Anion (Å)":                           anion_size,
            "Dielectric constant solvent":              diel,
            "Solvent":                                  solvent,
            "O/A ratio":                                oa,
            "Distribution ratio D":                     d_val,
            "Extraction efficiency E (%)":              e_val,
            "Partition coefficient (raw)":              p_val,
            "Notes / confidence":                       r.get("notes", ""),
        })

    df = pd.DataFrame(final, columns=COLUMNS)
    log.info(f"ASSEMBLER — {len(df)} rows")
    return df


# ── CELL 10 — Excel exporter ─────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
_THIN        = Side(style="thin", color="CCCCCC")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ROW_FILLS   = {
    "fig":   PatternFill("solid", fgColor="DDEEFF"),
    "table": PatternFill("solid", fgColor="E2EFDA"),
    "eq":    PatternFill("solid", fgColor="FFF3CD"),
    "text":  PatternFill("solid", fgColor="F5F5F5"),
}


def export_excel(df: pd.DataFrame, output_path: str) -> str:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extracted Data")

    wb = load_workbook(output_path)
    ws = wb.active
    ws.row_dimensions[1].height = 38

    for cell in ws[1]:
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.border    = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    src_col = list(df.columns).index("Figure / Table") + 1
    for row_idx in range(2, len(df) + 2):
        src  = str(ws.cell(row_idx, src_col).value or "").lower()
        key  = ("fig"   if "fig"   in src else
                "table" if "table" in src else
                "eq"    if "eq"    in src else "text")
        for col_idx in range(1, len(df.columns) + 1):
            c = ws.cell(row_idx, col_idx)
            c.fill      = _ROW_FILLS[key]
            c.border    = _BORDER
            c.alignment = Alignment(vertical="center", horizontal="center")

    for col in ws.columns:
        vals = [str(c.value or "") for c in col]
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(len(v) for v in vals) + 3, 42)
    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws_leg = wb.create_sheet("Legend")
    legend_rows = [
        ("Row colour",              "Source"),
        ("Blue",                    "Figure (vision-digitised)"),
        ("Green",                   "Table (LLM-parsed)"),
        ("Amber",                   "Equation (algebraic evaluation)"),
        ("Grey",                    "Text (value from running text)"),
        ("",                        ""),
        ("D ↔ E",                   "D = (E/100)·(O/A) / (1−E/100)  |  E = 100·D/(D+O/A)"),
        ("Partition coeff",         "Treated as equivalent to D; stored also in 'Partition coefficient (raw)'"),
        ("TBP vol%→mol/L",          f"mol/L = vol%×10×{TBP_DENSITY}/{TBP_MW}"),
        ("Anion conc",              "Set equal to acid concentration"),
        ("Ionic radii",             "From CATION_TABLE / ANION_TABLE (user-supplied)"),
        ("Dielectric const",        "From SOLVENT_TABLE (user-supplied)"),
        ("Initial metal conc",      "Concentration in the AQUEOUS phase BEFORE extraction (g/L)"),
        ("metal_override",          "Non-null when figure contains multiple metals; "
                                    "overrides primary metal from text agent"),
        ("Equation default range",  "If paper gives no range, defaults: "
                                    "TBP 5–100 vol% or 0.05–1.5 mol/L; acid 0.1–10 mol/L; pH 0–14"),
    ]
    for r_i, (k, v) in enumerate(legend_rows, 1):
        ws_leg.cell(r_i, 1, k).font = Font(bold=bool(k))
        ws_leg.cell(r_i, 2, v)
    ws_leg.column_dimensions["A"].width = 28
    ws_leg.column_dimensions["B"].width = 72

    wb.save(output_path)
    return str(Path(output_path).resolve())


# ── CELL 11 — RUN EVERYTHING ─────────────────────────────────────────────────

print("=" * 60)
print("Starting extraction pipeline...")
print("=" * 60)

# 1. Parse PDF
parsed = parse_pdf(PDF_PATH)

# 2. Text agent — metadata, equations, inline values
meta = run_text_agent(parsed)

# 3. Table agent
table_rows = run_table_agent(parsed)

# 4. Figure agent (vision)
figure_rows = run_figure_agent(parsed)

# 5. Equation evaluator
equation_rows = run_equation_agent(meta.get("equations", []), meta)

# 6. Text-stated data points
text_rows = []
for tp in meta.get("text_points", []):
    text_rows.append({
        "_source":         "Text",
        "_metal_override": None,
        "acid_mol_l":      tp.get("acid_mol_l"),
        "pH":              tp.get("pH"),
        "tbp_pct":         tp.get("tbp_pct"),
        "tbp_mol_l":       tp.get("tbp_mol_l") or _pct_to_mol(tp.get("tbp_pct")),
        "D":               tp.get("D"),
        "partition_coeff": tp.get("partition_coeff"),
        "E_pct":           tp.get("E_pct"),
        "metal_conc_g_l":  None,
        "oa_ratio":        None,
        "notes":           tp.get("notes", "from text"),
    })

all_rows = table_rows + figure_rows + equation_rows + text_rows

print(f"\nRows collected:")
print(f"  Tables    : {len(table_rows)}")
print(f"  Figures   : {len(figure_rows)}")
print(f"  Equations : {len(equation_rows)}")
print(f"  Text      : {len(text_rows)}")
print(f"  TOTAL     : {len(all_rows)}")

if not all_rows:
    log.warning("No data rows found — check PDF path and API key.")
    all_rows = [{
        "_source": "none", "_metal_override": None,
        "acid_mol_l": None, "pH": None, "tbp_pct": None,
        "tbp_mol_l": None, "D": None, "partition_coeff": None,
        "E_pct": None, "metal_conc_g_l": None,
        "oa_ratio": None, "notes": "no data found",
    }]

# 7. Assemble into DataFrame
df = assemble(all_rows, meta, Path(PDF_PATH).stem)

# 8. Export to Excel
out = export_excel(df, OUTPUT_PATH)

print(f"\n{'='*60}")
print(f"✅  Done — {len(df)} rows written to:")
print(f"   {out}")
print(f"{'='*60}")

# Preview
df